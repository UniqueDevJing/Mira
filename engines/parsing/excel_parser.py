"""Excel 解析: openpyxl 逐 sheet → markdown 表格块 + 结构化 tables。

依赖 openpyxl (lazy import: 未安装时仅 parse() 报错, 不影响模块导入与其他 parser)。
每个工作表转为一段 markdown 简表文本(前 50 行)便于检索命中, 完整矩阵存入 tables 供精确引用。
"""

import hashlib
import logging
from pathlib import Path

from engines.parsing.models import UIRDocument

logger = logging.getLogger(__name__)


class ExcelParser:
    supported_types = (".xlsx", ".xlsm")

    def parse(self, file_path: str) -> UIRDocument:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        blocks: list[dict] = []
        tables: list[dict] = []
        doc_id = hashlib.sha256(Path(file_path).read_bytes()).hexdigest()[:16]

        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [("" if c is None else str(c).strip()) for c in row]
                if any(cells):
                    rows.append(cells)
            if not rows:
                continue
            head = rows[0]
            body = rows[1 : min(len(rows), 51)]
            md = f"## 工作表：{ws.title}\n\n" + " | ".join(head)
            if body:
                md += "\n" + "\n".join(" | ".join(r) for r in body)
            blocks.append(
                {"type": "paragraph", "bbox": [], "content": md, "page_num": 1, "metadata": {"sheet": ws.title, "is_table": True}}
            )
            tables.append({"page_num": 1, "bbox": [], "matrix": rows, "headers": head})
        wb.close()

        update_time = int(Path(file_path).stat().st_mtime)
        return UIRDocument(
            doc_id=doc_id,
            source={"type": "excel", "path": file_path},
            pages=[{"page_num": 1, "blocks": blocks}],
            tables=tables,
            update_time=update_time,
        )
